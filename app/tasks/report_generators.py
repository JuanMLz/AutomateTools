# app/tasks/report_generators.py
import pandas as pd
import re
import unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from .utils import clean_program_name, slugify, get_weekday_key # Importa Utils

def _normalize(text):
    """Remove acentos e lowercase para comparações insensíveis."""
    nfkd = unicodedata.normalize('NFKD', str(text))
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

def generate_epg_from_simple_schedule(simple_schedule_df, epg_output_path):
    try:
        from .epg_database_manager import epg_manager
        
        df = simple_schedule_df.copy()
        df['inicio'] = pd.to_datetime(df['Data'] + ' ' + df['Horario'], format='%d/%m/%Y %H:%M')
        df = df.sort_values(by='inicio').reset_index(drop=True)
        
        # Carrega mapa Title → Unique ID do banco
        title_to_id = epg_manager.get_title_to_id_map()
        
        def resolve_unique_id(program_name):
            key = _normalize(program_name)
            if key in title_to_id:
                return title_to_id[key]
            return slugify(program_name)
            
        df['unique_id'] = df['Programa_Padronizado'].apply(resolve_unique_id)

        # Lógica visual EPG
        datas = sorted(df['inicio'].dt.date.unique())
        # The target format is exactly 'DD/MM/YYYY' for the column headers
        colunas_datas = [d.strftime('%d/%m/%Y') for d in datas]
        times_str = pd.date_range("00:00", "23:55", freq="5min").time.astype(str)
        indice_horarios = pd.to_datetime(times_str, format='%H:%M:%S')
        
        grade_df = pd.DataFrame(index=indice_horarios, columns=colunas_datas)
        grade_df.index.name = 'BRT'

        # Calcula horários de início arredondados e preenche a grade com suporte a cross-midnight
        for i, row in df.iterrows():
            dt_inicio = row['inicio']
            m = dt_inicio.minute
            m_round = 5 * round(m / 5)
            if m_round == 60:
                dt_inicio += pd.Timedelta(hours=1)
                dt_inicio = dt_inicio.replace(minute=0)
            else:
                dt_inicio = dt_inicio.replace(minute=m_round)
            
            # Fim é o início do próximo programa, ou +2h se for o último
            if i < len(df) - 1:
                dt_fim = df.iloc[i+1]['inicio']
                m_fim = dt_fim.minute
                m_fim_round = 5 * round(m_fim / 5)
                if m_fim_round == 60:
                    dt_fim += pd.Timedelta(hours=1)
                    dt_fim = dt_fim.replace(minute=0)
                else:
                    dt_fim = dt_fim.replace(minute=m_fim_round)
            else:
                dt_fim = dt_inicio + pd.Timedelta(hours=2)
                
            unique_id = row['unique_id']
            
            # Preenche os slots de 5 em 5 minutos de dt_inicio até dt_fim
            curr_time = dt_inicio
            while curr_time < dt_fim:
                data_str = curr_time.strftime('%d/%m/%Y')
                time_str = curr_time.strftime('%H:%M:%S')
                idx_inicio = pd.to_datetime(time_str, format='%H:%M:%S')
                
                if data_str in grade_df.columns and idx_inicio in grade_df.index:
                    grade_df.loc[idx_inicio, data_str] = unique_id
                
                curr_time += pd.Timedelta(minutes=5)

        # Carrega todo o Banco EPG (requisito do usuário: a aba deve ter tudo, não só a semana)
        df_epg_db = epg_manager.load_db()

        with pd.ExcelWriter(epg_output_path, engine='xlsxwriter', datetime_format='hh:mm:ss') as writer:
            grade_df.index = grade_df.index.strftime('%H:%M:%S')
            grade_df.to_excel(writer, sheet_name='Schedule')
            
            wb = writer.book
            ws = writer.sheets['Schedule']
            merge_fmt = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
            ws.set_column('A:A', 10)
            ws.set_column('B:Z', 25)

            # Mescla as células contíguas do mesmo programa
            for col_num, _ in enumerate(grade_df.columns):
                excel_col = col_num + 1
                
                # Inicia no primeiro valor da coluna
                start_row = 1  # Considerando que row_num = 0 vira excel_row = 1
                last_txt = grade_df.iat[0, col_num]
                if pd.isna(last_txt): 
                    last_txt = ""
                else:
                    last_txt = str(last_txt)

                for row_num in range(1, len(grade_df)):
                    excel_row = row_num + 1
                    val = grade_df.iat[row_num, col_num]
                    if pd.isna(val): 
                        val = ""
                    else:
                        val = str(val)
                    
                    if val != last_txt:
                        # O valor mudou, escreve o bloco do programa anterior (se não for vazio)
                        if last_txt != "":
                            end_row = excel_row - 1
                            if end_row > start_row: 
                                ws.merge_range(start_row, excel_col, end_row, excel_col, last_txt, merge_fmt)
                            else: 
                                ws.write(start_row, excel_col, last_txt, merge_fmt)
                        
                        start_row = excel_row
                        last_txt = val
                
                # Grava o último bloco da coluna ao terminar o laço
                if last_txt != "":
                    end_row = len(grade_df)
                    if end_row > start_row: 
                        ws.merge_range(start_row, excel_col, end_row, excel_col, last_txt, merge_fmt)
                    else: 
                        ws.write(start_row, excel_col, last_txt, merge_fmt)
            
            # 5. Salva a Aba EPG Inteira
            df_epg_db.to_excel(writer, sheet_name='EPG', index=False)

        return f"Sucesso! EPG salvo em '{epg_output_path}'"

    except Exception as e:
        import traceback
        return f"Erro EPG: {e} | {traceback.format_exc()}"

def generate_comparison_report(clean_schedule_df, excel_anterior_path, output_path):
    try:
        df_novo = clean_schedule_df.copy()

        # 1. Leitura do Antigo
        try:
            df_antigo = pd.read_excel(excel_anterior_path, header=0)
            if 'Data' not in df_antigo.columns: raise ValueError()
        except:
            df_antigo = pd.read_excel(excel_anterior_path, header=2)

        df_antigo = df_antigo.loc[:, ~df_antigo.columns.str.contains('^Unnamed')]
        df_antigo.columns = df_antigo.columns.str.strip()
        if 'Programa' in df_antigo.columns:
            df_antigo.rename(columns={'Programa': 'Programa_Padronizado'}, inplace=True)

        # === CORREÇÃO 3: Aplica a MESMA limpeza do Utils no arquivo antigo ===
        # Isso garante que se o Excel antigo tiver "Nome " (espaço no fim), ele limpa.
        if 'Programa_Padronizado' in df_antigo.columns:
            df_antigo['Programa_Padronizado'] = df_antigo['Programa_Padronizado'].apply(clean_program_name)

        df_antigo['Data'] = df_antigo['Data'].astype(str)
        # Normaliza Hora Antiga
        def _norm_time_old(val):
            if pd.isna(val): return "00:00"
            if hasattr(val, 'hour'): return f"{int(val.hour):02}:{int(val.minute):02}"
            s = str(val).strip()
            m = re.search(r'(\d{1,2}:\d{2})', s)
            return m.group(1) if m else s[:5]
        df_antigo['Horario'] = df_antigo['Horario'].apply(_norm_time_old)

        # Gera chaves (usa utils)
        df_novo['chave'] = df_novo.apply(lambda row: get_weekday_key(row), axis=1)
        df_antigo['chave'] = df_antigo.apply(lambda row: get_weekday_key(row), axis=1)

        mapa_antigo = df_antigo.drop_duplicates(subset=['chave'], keep='last').set_index('chave')['Programa_Padronizado'].to_dict()
        db_metadados = df_antigo.drop_duplicates(subset=['Programa_Padronizado'], keep='last').set_index('Programa_Padronizado').to_dict('index')

        colunas_principais = {'Data', 'Horario', 'Programa_Padronizado', 'chave', 'Status', 'Programa_Bruto'}
        colunas_extras = [c for c in df_antigo.columns if c not in colunas_principais]

        registros = []
        for _, row in df_novo.iterrows():
            chave = row['chave']
            # Limpeza garantida na comparação
            prog_novo = clean_program_name(row['Programa_Padronizado'])
            prog_antigo = mapa_antigo.get(chave)
            
            if not prog_antigo or str(prog_antigo) == 'nan': status = 'NOVO'
            elif prog_novo != prog_antigo: status = 'ALTERADO'
            else: status = 'SEM MUDANÇA'
            
            item = {'Data': row['Data'], 'Horario': row['Horario'], 'Programa': prog_novo, 'Status': status}
            meta = db_metadados.get(prog_novo, {})
            for col in colunas_extras: item[col] = meta.get(col, "") if pd.notna(meta.get(col)) else ""
            registros.append(item)

        # 3. Escrita (Mantida igual)
        wb = load_workbook(excel_anterior_path)
        ws = wb.active
        fill_green = PatternFill("solid", fgColor="C6EFCE")
        fill_yellow = PatternFill("solid", fgColor="FFFF00")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal='left', vertical='center')

        ws.delete_rows(4, amount=(ws.max_row + 100))
        curr_row = 4
        last_date = None
        cols_order = ['Data', 'Horario', 'Programa'] + colunas_extras

        for reg in registros:
            if last_date and reg['Data'] != last_date:
                for c in range(2, len(cols_order) + 3):
                    cell = ws.cell(curr_row, c); cell.fill = fill_yellow; cell.border = border
                curr_row += 1
            last_date = reg['Data']
            is_changed = reg['Status'] in ['NOVO', 'ALTERADO']
            for i, col in enumerate(cols_order):
                cell = ws.cell(curr_row, i + 2); cell.value = reg.get(col, ""); cell.border = border; cell.alignment = align
                if is_changed: cell.fill = fill_green
            curr_row += 1

        wb.save(output_path)
        return f"Sucesso! Salvo em '{output_path}'"

    except Exception as e:
        import traceback
        return f"Erro Comparação: {e} | {traceback.format_exc()}"