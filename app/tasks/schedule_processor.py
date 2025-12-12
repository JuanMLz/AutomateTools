# app/tasks/schedule_processor.py
"""
Módulo responsável por toda a lógica de processamento de grades:
1. Extração de dados de PDFs.
2. Limpeza e Padronização.
3. Geração de EPG Visual.
4. Geração de Relatórios Comparativos (Excel Template).
"""

import re
import unicodedata
import fitz  # PyMuPDF
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from .mapping_manager import mapping_manager

# ======================================================================
# == 1. FUNÇÕES AUXILIARES (Internas)                                 ==
# ======================================================================

def _slugify(text):
    """Converte texto para formato URL amigável (ex: 'Mãe Maria' -> 'mae-maria')."""
    if not text: return ""
    text = unicodedata.normalize('NFKD', str(text)).encode('ascii', 'ignore').decode('utf-8')
    text = text.lower()
    text = re.sub(r'[^a-z0-9]+', '-', text).strip('-')
    return text

def _extract_date_from_pdf(pdf_path):
    """Procura a primeira data no formato DD/MM/AAAA dentro do PDF."""
    try:
        doc = fitz.open(pdf_path)
        for page in doc:
            text = page.get_text()
            match = re.search(r'\d{2}/\d{2}/\d{4}', text)
            if match: return match.group(0)
        return ""
    except Exception:
        return ""

def _get_weekday_key(row):
    """
    Gera chave baseada em Dia da Semana Real (0=seg, 1=ter, ..., 6=dom) + HH:MM.
    Usa datetime.weekday() para garantir consistência independente da ordem das datas.
    Aceita 'Data' como string (DD/MM/YYYY) ou objeto datetime.
    """
    try:
        # Normaliza horário para HH:MM
        horario_raw = row.get('Horario', "")
        if hasattr(horario_raw, 'hour') and hasattr(horario_raw, 'minute'):
            time_str = f"{int(horario_raw.hour):02}:{int(horario_raw.minute):02}"
        else:
            s = str(horario_raw).strip()
            m = re.search(r'(\d{1,2}:\d{2})', s)
            time_str = m.group(1) if m else s[:5]
        
        # Converte data para datetime e extrai weekday (0=seg, ..., 6=dom)
        data_raw = row.get('Data', '')
        if isinstance(data_raw, str):
            try:
                data_obj = pd.to_datetime(data_raw, format='%d/%m/%Y')
            except:
                data_obj = pd.to_datetime(data_raw)
        else:
            data_obj = pd.to_datetime(data_raw)
        
        weekday = data_obj.weekday()  # 0=seg, 1=ter, ..., 6=dom
        return f"{weekday}_{time_str}"
    except Exception:
        return f"ERR_{row.get('Data')}_{row.get('Horario')}"

def _extract_raw_data_from_pdfs(pdf_paths):
    """Lê as coordenadas X/Y do PDF para separar Horário de Programa."""
    all_schedule_data = []
    COLUMN_DIVIDER_X = 70.0 # Divisor visual entre coluna de hora e nome

    for pdf_path in pdf_paths:
        date = _extract_date_from_pdf(pdf_path)
        doc = fitz.open(pdf_path)
        page = doc[0] # Assume que a grade está na primeira página
        words = page.get_text("words")
        
        # Agrupa palavras por linha (eixo Y)
        lines = {}
        for word in words:
            y0 = word[1]
            line_key = int(y0 // 10) # Agrupamento aproximado
            if line_key not in lines: lines[line_key] = []
            lines[line_key].append(word)
        
        # Processa cada linha
        for line_key in sorted(lines.keys()):
            line_words = sorted(lines[line_key], key=lambda w: w[0])
            horario = ""
            programa_parts = []
            
            for word in line_words:
                if word[0] < COLUMN_DIVIDER_X:
                    horario = word[4]
                else:
                    programa_parts.append(word[4])
            
            # Só adiciona se tiver horário válido
            if horario and horario[:1].isdigit():
                all_schedule_data.append({
                    'Data': date,
                    'Horario': horario,
                    'Programa_Bruto': " ".join(programa_parts)
                })
                
    return pd.DataFrame(all_schedule_data)

def find_unmapped_programs(pdf_paths=None, df_extracted=None):
    """
    Retorna (unmapped_list, None) ou (None, mensagem_de_erro).
    Aceita `df_extracted` (DataFrame) ou `pdf_paths` (lista).
    Lida com DataFrames que contenham 'Programa_Bruto' OU 'Programa_Padronizado'.
    Retorna nomes originais (brutos quando disponíveis) para exibição no editor.
    """
    mapping_dict, err = mapping_manager.load_mapping_as_dict()
    if err:
        return None, err

    try:
        # Obter dataframe de origem
        if df_extracted is None:
            if not pdf_paths:
                return [], None
            df_raw = _extract_raw_data_from_pdfs(pdf_paths)
        else:
            df_raw = df_extracted.copy()

        if df_raw is None or df_raw.empty:
            return [], None

        # Normalização leve: remove acentos, espaços extras e lower
        def _norm(s):
            if s is None:
                return ""
            s = str(s).strip()
            s = unicodedata.normalize('NFKD', s).encode('ascii', 'ignore').decode('utf-8')
            s = re.sub(r'\s+', ' ', s)
            return s.lower()

        # Dois cenários:
        # A) Temos 'Programa_Bruto' => comparamos contra as chaves (Nome_do_PDF) do mapping
        if 'Programa_Bruto' in df_raw.columns:
            mapped_keys_norm = { _norm(k) for k in mapping_dict.keys() if pd.notna(k) }
            unique_raw = pd.Series(df_raw['Programa_Bruto'].astype(str).unique())
            unmapped = [raw for raw in unique_raw if _norm(raw) not in mapped_keys_norm]
            return unmapped, None

        # B) Temos apenas 'Programa_Padronizado' (output já com o replace aplicado)
        #    Nesse caso, assumimos que valores que NÃO aparecem em mapping.values() são não-mapeados.
        if 'Programa_Padronizado' in df_raw.columns:
            mapped_values_norm = { _norm(v) for v in mapping_dict.values() if pd.notna(v) }
            unique_pad = pd.Series(df_raw['Programa_Padronizado'].astype(str).unique())
            # itens cujo padronizado não aparece entre valores mapeados => provavelmente não mapeados
            unmapped = [p for p in unique_pad if _norm(p) not in mapped_values_norm]
            return unmapped, None

        # Se nenhuma coluna esperada existir, devolve erro explicativo
        return None, "Erro ao detectar programas não mapeados: colunas esperadas ausentes ('Programa_Bruto' ou 'Programa_Padronizado')."

    except Exception as e:
        return None, f"Erro ao detectar programas não mapeados: {e}"
# ======================================================================
# == 2. FUNÇÕES PRINCIPAIS (Tasks)                                    ==
# ======================================================================

def extract_and_clean_from_pdfs(pdf_paths):
    """Extrai dados, ordena cronologicamente e aplica o De-Para (Blindado)."""
    
    # 1. Carrega o dicionário
    raw_mapping, error = mapping_manager.load_mapping_as_dict()
    if error: return None, error

    # BLINDAGEM DO MAPEAMENTO: Remove espaços das chaves do dicionário
    # Ex: " Programa X " vira "Programa X" no índice de busca
    mapping_dict = {k.strip(): v for k, v in raw_mapping.items()}

    try:
        df_extracted = _extract_raw_data_from_pdfs(pdf_paths)
        if df_extracted.empty: return None, "Erro: PDFs vazios ou ilegíveis."

        # 2. Ordenação Cronológica (Mantida igual)
        df_extracted['temp_data'] = pd.to_datetime(df_extracted['Data'], format='%d/%m/%Y', errors='coerce')
        df_extracted['temp_hora_dt'] = pd.to_datetime(df_extracted['Horario'], format='%H:%M', errors='coerce')
        mask_na = df_extracted['temp_hora_dt'].isna()
        if mask_na.any():
            df_extracted.loc[mask_na, 'temp_hora_dt'] = pd.to_datetime(df_extracted.loc[mask_na, 'Horario'], format='%H:%M:%S', errors='coerce')

        df_extracted['temp_hora'] = df_extracted['temp_hora_dt'].dt.time
        df_extracted.sort_values(by=['temp_data', 'temp_hora'], inplace=True)

        df_extracted['Horario'] = df_extracted['temp_hora'].apply(lambda t: f"{t.hour:02}:{t.minute:02}" if pd.notna(t) else "")
        df_extracted.drop(columns=['temp_hora_dt', 'temp_hora', 'temp_data'], inplace=True, errors='ignore')

        # 3. BLINDAGEM DA EXTRAÇÃO: Limpeza rigorosa do nome bruto
        # Remove espaços do início/fim e converte para string
        df_extracted['Programa_Bruto'] = df_extracted['Programa_Bruto'].astype(str).str.strip()
        
        # Remove espaços duplos no meio do nome (ex: "Jornal  Hoje" -> "Jornal Hoje")
        df_extracted['Programa_Bruto'] = df_extracted['Programa_Bruto'].str.replace(r'\s+', ' ', regex=True)

        # 4. Aplica Mapeamento (Agora com dados limpos vs chaves limpas)
        df_extracted['Programa_Padronizado'] = df_extracted['Programa_Bruto'].replace(mapping_dict)
        
        # Gera chave
        df_extracted['chave'] = df_extracted.apply(lambda row: _get_weekday_key(row), axis=1)
        
        return df_extracted[['Data', 'Horario', 'Programa_Bruto', 'Programa_Padronizado', 'chave']], None
    except Exception as e:
        import traceback
        return None, f"Erro na extração: {e} | {traceback.format_exc()}"

def generate_epg_from_simple_schedule(simple_schedule_df, epg_output_path):
    """
    Gera o Excel Visual (EPG) com células mescladas e aba de Database.
    INTEGRAÇÃO: Atualiza o epg_database.csv com novos títulos encontrados.
    """
    try:
        # Importação sob demanda para evitar ciclos
        from .epg_database_manager import epg_manager
        
        df = simple_schedule_df.copy()
        
        # 1. Preparação de Datas e Horários
        df['inicio'] = pd.to_datetime(df['Data'] + ' ' + df['Horario'], format='%d/%m/%Y %H:%M')
        df['titulo_slug'] = df['Programa_Padronizado'].apply(_slugify)
        df = df.sort_values(by='inicio').reset_index(drop=True)
        
        # === ATUALIZAÇÃO DO BANCO DE DADOS ===
        # Extrai lista de títulos e slugs para verificar se são novos
        # Usamos drop_duplicates para processar cada programa apenas uma vez
        unique_progs = df[['titulo_slug', 'Programa_Padronizado']].drop_duplicates()
        slugs_list = unique_progs['titulo_slug'].tolist()
        titles_list = unique_progs['Programa_Padronizado'].tolist()
        
        added_count = epg_manager.update_with_new_programs(slugs_list, titles_list)
        # =====================================

        datas = sorted(df['inicio'].dt.date.unique())
        colunas_datas = [d.strftime('%d/%m/%Y') for d in datas]
        
        # Cria índice de horários (00:00 a 23:55)
        times_str = pd.date_range("00:00", "23:55", freq="5min").time.astype(str)
        indice_horarios = pd.to_datetime(times_str, format='%H:%M:%S')
        
        grade_df = pd.DataFrame(index=indice_horarios, columns=colunas_datas)
        grade_df.index.name = 'BRT'

        # Preenche a Grade Visual
        for _, row in df.iterrows():
            data_str = row['inicio'].strftime('%d/%m/%Y')
            h, m = row['inicio'].hour, row['inicio'].minute
            m_round = 5 * round(m / 5)
            if m_round == 60: h += 1; m_round = 0; 
            if h == 24: h = 0
            
            time_str = f"{h:02}:{m_round:02}:00"
            idx_inicio = pd.to_datetime(time_str, format='%H:%M:%S')
            
            if data_str in grade_df.columns and idx_inicio in grade_df.index:
                grade_df.loc[idx_inicio, data_str] = row['titulo_slug']

        # Carrega o banco atualizado para salvar na aba 2
        df_epg_db = epg_manager.load_db()

        # Escrita com XlsxWriter (2 Abas)
        with pd.ExcelWriter(epg_output_path, engine='xlsxwriter') as writer:
            
            # --- ABA 1: SCHEDULE ---
            grade_df.index = grade_df.index.strftime('%H:%M')
            grade_df.to_excel(writer, sheet_name='Schedule')
            
            wb = writer.book
            ws = writer.sheets['Schedule']
            merge_fmt = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
            
            ws.set_column('A:A', 10)
            ws.set_column('B:Z', 25)

            # Algoritmo de Mesclagem
            for col_num, _ in enumerate(grade_df.columns):
                excel_col = col_num + 1
                start_row = -1
                last_txt = None
                
                for row_num in range(len(grade_df)):
                    excel_row = row_num + 1
                    val = grade_df.iloc[row_num, col_num]
                    
                    if pd.notna(val) and val != "":
                        if start_row != -1:
                            end_row = excel_row - 1
                            if end_row > start_row:
                                ws.merge_range(start_row, excel_col, end_row, excel_col, last_txt, merge_fmt)
                            else:
                                ws.write(start_row, excel_col, last_txt, merge_fmt)
                        start_row = excel_row
                        last_txt = val
                    
                    if row_num == len(grade_df) - 1 and start_row != -1:
                        if excel_row > start_row:
                            ws.merge_range(start_row, excel_col, excel_row, excel_col, last_txt, merge_fmt)
                        else:
                            ws.write(start_row, excel_col, last_txt, merge_fmt)
            
            # --- ABA 2: EPG (DATABASE) ---
            df_epg_db.to_excel(writer, sheet_name='EPG', index=False)
            ws_epg = writer.sheets['EPG']
            ws_epg.set_column('A:A', 25) # Unique ID
            ws_epg.set_column('B:B', 30) # Title
            ws_epg.set_column('H:H', 40) # Short Desc
            ws_epg.set_column('I:I', 60) # Long Desc

        msg_extra = f" (+{added_count} novos programas cadastrados)" if added_count > 0 else ""
        return f"Sucesso! EPG salvo em '{epg_output_path}'{msg_extra}"

    except Exception as e:
        import traceback
        return f"Erro EPG: {e} | {traceback.format_exc()}"

def generate_comparison_report(clean_schedule_df, excel_anterior_path, output_path):
    """
    Gera relatório comparativo usando a planilha anterior como Template.
    BLINDAGEM: Normaliza strings (remove espaços) antes de comparar.
    """
    try:
        df_novo = clean_schedule_df.copy()

        # 1. Leitura Inteligente do Template
        try:
            df_antigo = pd.read_excel(excel_anterior_path, header=0)
            if 'Data' not in df_antigo.columns: raise ValueError()
        except:
            df_antigo = pd.read_excel(excel_anterior_path, header=2)

        # Limpeza de colunas do Template
        df_antigo = df_antigo.loc[:, ~df_antigo.columns.str.contains('^Unnamed')]
        df_antigo.columns = df_antigo.columns.str.strip()
        
        if 'Programa' in df_antigo.columns:
            df_antigo.rename(columns={'Programa': 'Programa_Padronizado'}, inplace=True)

        # === BLINDAGEM DO TEMPLATE (Limpeza de Strings) ===
        # Garante que os nomes no Excel antigo não tenham espaços invisíveis
        if 'Programa_Padronizado' in df_antigo.columns:
            df_antigo['Programa_Padronizado'] = df_antigo['Programa_Padronizado'].astype(str).str.strip()
            df_antigo['Programa_Padronizado'] = df_antigo['Programa_Padronizado'].str.replace(r'\s+', ' ', regex=True)
        # ==================================================

        # Normaliza Data/Hora
        df_antigo['Data'] = df_antigo['Data'].astype(str)
        
        def _normalize_time(val):
            if pd.isna(val): return "00:00"
            if hasattr(val, 'hour'): return f"{int(val.hour):02}:{int(val.minute):02}"
            s = str(val).strip()
            m = re.search(r'(\d{1,2}:\d{2})', s)
            return m.group(1) if m else s[:5]
        
        df_antigo['Horario'] = df_antigo['Horario'].apply(_normalize_time)
        
        # Gera chaves
        df_novo['chave'] = df_novo.apply(lambda row: _get_weekday_key(row), axis=1)
        df_antigo['chave'] = df_antigo.apply(lambda row: _get_weekday_key(row), axis=1)

        # Cria mapas de referência
        # Pega a última ocorrência para evitar duplicatas de chave
        df_antigo_unique = df_antigo.drop_duplicates(subset=['chave'], keep='last')
        mapa_antigo = df_antigo_unique.set_index('chave')['Programa_Padronizado'].to_dict()
        
        # Mapa de Metadados (Sinopses, etc.)
        df_meta_unique = df_antigo.drop_duplicates(subset=['Programa_Padronizado'], keep='last')
        db_metadados = df_meta_unique.set_index('Programa_Padronizado').to_dict('index')

        colunas_principais = {'Data', 'Horario', 'Programa_Padronizado', 'chave', 'Status', 'Programa_Bruto'}
        colunas_extras = [c for c in df_antigo.columns if c not in colunas_principais]

        # 2. Comparação
        registros = []
        for _, row in df_novo.iterrows():
            chave = row['chave']
            # Garante limpeza também no nome novo (vindo da extração)
            prog_novo = str(row['Programa_Padronizado']).strip().replace('  ', ' ')
            
            prog_antigo = mapa_antigo.get(chave)
            
            # Lógica de Status
            if not prog_antigo or str(prog_antigo) == 'nan':
                status = 'NOVO'
            elif prog_novo != prog_antigo:
                status = 'ALTERADO'
            else:
                status = 'SEM MUDANÇA'
            
            item = {
                'Data': row['Data'], 'Horario': row['Horario'], 'Programa': prog_novo, 'Status': status
            }
            
            # Recupera Metadados (usando o nome limpo como chave)
            meta_row = db_metadados.get(prog_novo, {})
            for col in colunas_extras:
                val = meta_row.get(col, "")
                item[col] = val if pd.notna(val) else ""
            
            registros.append(item)

        # 3. Escrita no Excel (Lógica Visual - Mantida igual)
        wb = load_workbook(excel_anterior_path)
        ws = wb.active
        
        fill_green = PatternFill("solid", fgColor="C6EFCE")
        fill_yellow = PatternFill("solid", fgColor="FFFF00")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align = Alignment(horizontal='left', vertical='center')

        start_row = 4
        ws.delete_rows(start_row, amount=(ws.max_row + 100))

        curr_row = start_row
        last_date = None
        cols_order = ['Data', 'Horario', 'Programa'] + colunas_extras

        for reg in registros:
            # Linha Amarela
            if last_date and reg['Data'] != last_date:
                for c in range(2, len(cols_order) + 3):
                    cell = ws.cell(curr_row, c)
                    cell.fill = fill_yellow
                    cell.border = border
                ws.row_dimensions[curr_row].height = 15
                curr_row += 1
            
            last_date = reg['Data']
            is_changed = reg['Status'] in ['NOVO', 'ALTERADO']

            for i, col in enumerate(cols_order):
                cell = ws.cell(curr_row, i + 2)
                cell.value = reg.get(col, "")
                cell.border = border
                cell.alignment = align
                if is_changed: cell.fill = fill_green
            
            curr_row += 1

        wb.save(output_path)
        return f"Sucesso! Salvo em '{output_path}'"

    except Exception as e:
        import traceback
        return f"Erro Comparação: {e} | {traceback.format_exc()}"